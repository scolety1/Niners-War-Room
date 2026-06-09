from __future__ import annotations

# ruff: noqa: E501
import csv
import re
from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

PACKAGE_ROOT = Path(
    "local_exports/league_history/drafts/incoming/"
    "draft_prep_codex_package/draft_prep_codex_package"
)
OUTPUT_ROOT = Path("local_exports/model_v4/draft_prep/latest")
DOC_ROOT = Path("docs/model_v4")
ACTIVE_PACK = Path("local_exports/data_packs/lve_sleeper_20260505_pdf_ranks")
DRAFT_POOL_PREVIEW_PACK = Path(
    "local_exports/data_packs/lve_sleeper_20260505_pdf_ranks_draft_pool_20260508_213233"
)

EXPLICIT_2025_USER_DRAFTED = {
    "kaleb johnson",
    "jayden higgins",
    "luther burden iii",
    "dylan sampson",
    "trevor etienne",
}

OWNED_PICK_LABELS = ("2026 1.03", "2026 1.04", "2026 2.04", "2026 2.08", "2026 5.04")

PRIOR_HISTORY_HEADER = (
    "source_file",
    "source_type",
    "draft_year",
    "row_source_area",
    "round",
    "pick",
    "overall_pick",
    "draft_slot_label",
    "expected_slot",
    "drafted_at",
    "user_rank",
    "drafting_team",
    "drafting_manager",
    "player",
    "position",
    "nfl_team_at_draft",
    "college_team",
    "draft_capital",
    "age",
    "user_note",
    "user_drafted_flag",
    "user_must_draft_at_cost_flag",
    "highlight_color_context",
    "player_type",
    "normalized_player_key",
    "transcription_confidence",
    "allowed_use",
    "blocked_use",
    "warning_flags",
    "data_needed",
)

BEHAVIOR_HEADER = (
    "draft_year",
    "round",
    "picks_count",
    "rookie_count",
    "veteran_or_free_agent_count",
    "qb_count",
    "rb_count",
    "wr_count",
    "te_count",
    "k_count",
    "average_age",
    "first_round_veteran_count",
    "first_round_rookie_count",
    "notes",
)

READINESS_HEADER = (
    "source_area",
    "source_path",
    "exists",
    "rows",
    "readiness_status",
    "allowed_use",
    "blocked_use",
    "notes",
)

POOL_HEADER = (
    "player",
    "position",
    "nfl_team",
    "college_team",
    "age",
    "source_type",
    "draftable_status",
    "legal_draftable",
    "roster_owner",
    "protected_status",
    "rookie_status",
    "free_agent_status",
    "dropped_veteran_status",
    "manual_status",
    "nwr_draft_value",
    "nwr_rookie_score",
    "nwr_dynasty_score",
    "expected_league_slot_context",
    "nwr_slot_context",
    "value_threshold_context",
    "prospect_talent_context",
    "landing_spot_context",
    "draft_capital_context",
    "role_path_context",
    "trust_status",
    "warning_flags",
    "data_needed",
    "allowed_use",
    "blocked_use",
    "source_path",
    "source_column",
    "lineage_class",
)


@dataclass(frozen=True)
class DraftPrepBuildResult:
    prior_history_rows: int
    behavior_rows: int
    readiness_rows: int
    scouting_pool_rows: int
    confirmed_legal_pool_ready: bool
    scouting_pool_ready: bool
    output_root: Path


def build_draft_prep_data_foundation(
    *,
    package_root: str | Path = PACKAGE_ROOT,
    output_root: str | Path = OUTPUT_ROOT,
    doc_root: str | Path = DOC_ROOT,
) -> DraftPrepBuildResult:
    package = Path(package_root)
    output = Path(output_root)
    docs = Path(doc_root)
    output.mkdir(parents=True, exist_ok=True)
    docs.mkdir(parents=True, exist_ok=True)

    prior_rows = normalize_prior_draft_history(package)
    _write_csv(output / "prior_league_draft_history_review_rows.csv", PRIOR_HISTORY_HEADER, prior_rows)

    behavior_rows = build_league_behavior_summary(prior_rows)
    _write_csv(output / "prior_league_draft_behavior_summary.csv", BEHAVIOR_HEADER, behavior_rows)

    readiness_rows = build_draftable_pool_source_readiness()
    _write_csv(output / "draftable_pool_source_readiness.csv", READINESS_HEADER, readiness_rows)

    scouting_rows = build_scouting_prep_pool_rows()
    _write_csv(output / "scouting_prep_pool_review_rows.csv", POOL_HEADER, scouting_rows)

    _write_current_state_audit(docs / "DRAFT_PREP_CURRENT_STATE_AUDIT_20260609.md", readiness_rows)
    _write_prior_normalization_doc(
        docs / "PRIOR_DRAFT_HISTORY_NORMALIZATION_20260609.md",
        prior_rows,
        package,
    )
    _write_behavior_doc(
        docs / "PRIOR_LEAGUE_DRAFT_BEHAVIOR_SUMMARY_20260609.md",
        behavior_rows,
        prior_rows,
    )
    _write_source_contract(docs / "DRAFTABLE_POOL_SOURCE_CONTRACT_20260609.md")
    _write_readiness_doc(
        docs / "DRAFTABLE_POOL_SOURCE_READINESS_20260609.md",
        readiness_rows,
        scouting_rows,
    )
    _write_pick_window_spec(docs / "DRAFT_PREP_PICK_WINDOW_SPEC_20260609.md")
    _write_page_architecture(docs / "DRAFT_PREP_PAGE_ARCHITECTURE_20260609.md")

    legal_ready = _source_ready(readiness_rows, "active confirmed legal draftable pool")
    scouting_ready = bool(scouting_rows)
    return DraftPrepBuildResult(
        prior_history_rows=len(prior_rows),
        behavior_rows=len(behavior_rows),
        readiness_rows=len(readiness_rows),
        scouting_pool_rows=len(scouting_rows),
        confirmed_legal_pool_ready=legal_ready,
        scouting_pool_ready=scouting_ready,
        output_root=output,
    )


def normalize_prior_draft_history(package_root: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    raw = package_root / "raw_uploaded_files"
    for path in sorted(raw.glob("*.xlsx")):
        if path.name.startswith("2025"):
            rows.extend(_normalize_2025_workbook(path))
        elif path.name.startswith("2024"):
            rows.extend(_normalize_2024_workbook(path))
        else:
            rows.extend(_normalize_generic_workbook(path))
    pdf_csv = package_root / "drafts_pdf_official_transcription_best_effort.csv"
    if pdf_csv.exists():
        rows.extend(_normalize_pdf_transcription(pdf_csv))
    return rows


def build_league_behavior_summary(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    for row in _deduped_behavior_rows(rows):
        if not _include_in_behavior_summary(row):
            continue
        if not row.get("round"):
            continue
        grouped[(str(row.get("draft_year") or ""), str(row.get("round") or ""))].append(row)

    out: list[dict[str, object]] = []
    for (year, round_label), items in sorted(grouped.items(), key=lambda pair: (pair[0][0], _safe_int(pair[0][1]))):
        positions = Counter(str(item.get("position") or "").upper() for item in items)
        ages = [_safe_float(item.get("age")) for item in items if _safe_float(item.get("age")) is not None]
        rookie_count = sum(1 for item in items if item.get("player_type") == "rookie")
        veteran_count = sum(1 for item in items if item.get("player_type") == "veteran_or_free_agent")
        first_round = str(round_label) == "1"
        low_conf = sum(1 for item in items if item.get("transcription_confidence") == "low")
        out.append(
            {
                "draft_year": year,
                "round": round_label,
                "picks_count": len(items),
                "rookie_count": rookie_count,
                "veteran_or_free_agent_count": veteran_count,
                "qb_count": positions.get("QB", 0),
                "rb_count": positions.get("RB", 0),
                "wr_count": positions.get("WR", 0),
                "te_count": positions.get("TE", 0),
                "k_count": positions.get("K", 0),
                "average_age": round(sum(ages) / len(ages), 2) if ages else "",
                "first_round_veteran_count": veteran_count if first_round else 0,
                "first_round_rookie_count": rookie_count if first_round else 0,
                "notes": (
                    f"{low_conf} low-confidence PDF row(s); position/player-type counts incomplete."
                    if low_conf
                    else "Context-only summary; not a model input."
                ),
            }
        )
    return out


def _deduped_behavior_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    by_slot: dict[tuple[str, str], dict[str, object]] = {}
    no_slot: list[dict[str, object]] = []
    for row in rows:
        if not _include_in_behavior_summary(row):
            continue
        year = str(row.get("draft_year") or "")
        round_label = str(row.get("round") or "")
        pick = str(row.get("pick") or "")
        overall = str(row.get("overall_pick") or "")
        slot = f"{round_label}.{pick}" if round_label and pick else overall
        if not year or not slot:
            no_slot.append(row)
            continue
        key = (year, slot)
        current = by_slot.get(key)
        if current is None or _behavior_preference(row) > _behavior_preference(current):
            by_slot[key] = row
    return [*by_slot.values(), *no_slot]


def _behavior_preference(row: dict[str, object]) -> tuple[int, int]:
    source_area = str(row.get("row_source_area") or "")
    confidence = str(row.get("transcription_confidence") or "")
    structured_score = 2 if source_area == "user_rank_board" else 1
    confidence_score = {"high": 3, "medium": 2, "low": 1}.get(confidence, 2)
    return structured_score, confidence_score


def _include_in_behavior_summary(row: dict[str, object]) -> bool:
    source_area = str(row.get("row_source_area") or "")
    if source_area == "actual_draft":
        return True
    if source_area == "user_rank_board" and str(row.get("source_file") or "").startswith("2025"):
        return bool(row.get("drafted_at"))
    return False


def build_draftable_pool_source_readiness() -> list[dict[str, object]]:
    source_specs = (
        (
            "active confirmed legal draftable pool",
            ACTIVE_PACK / "fact_rookie_draftables.csv",
            False,
            "would support app-facing legal rookies if active pack included it",
            "not present in active pack; do not infer missing means no rookies",
        ),
        (
            "active available veterans/free agents",
            ACTIVE_PACK / "fact_available_veterans.csv",
            False,
            "would support app-facing legal free agents if active pack included it",
            "not present in active pack; do not infer no veterans/free agents",
        ),
        (
            "active manual draftables",
            ACTIVE_PACK / "fact_manual_draftables.csv",
            False,
            "manual additions if supplied",
            "not present in active pack",
        ),
        (
            "protected roster source",
            ACTIVE_PACK / "fact_rosters.csv",
            True,
            "protected/unavailable roster context",
            "must not be used to mark players draftable",
        ),
        (
            "preview rookie draftables",
            DRAFT_POOL_PREVIEW_PACK / "fact_rookie_draftables.csv",
            False,
            "scouting prep and draftable-source review",
            "not promoted to active pack by this task",
        ),
        (
            "preview free agents",
            DRAFT_POOL_PREVIEW_PACK / "fact_available_veterans.csv",
            False,
            "scouting prep and free-agent review",
            "not promoted to active pack by this task",
        ),
        (
            "preview manual draftables",
            DRAFT_POOL_PREVIEW_PACK / "fact_manual_draftables.csv",
            False,
            "manual source review if rows exist",
            "empty file is not proof no manual adds are needed",
        ),
        (
            "dropped/released veterans",
            Path("local_exports/league_history/roster_declaration/latest/dropped_veterans.csv"),
            True,
            "confirmed post-declaration legal veteran pool",
            "missing until Roster Declaration Day source is supplied",
        ),
        (
            "Model v4 rookie review",
            Path("local_exports/model_v4/rookie_draft_review/latest/rookie_draft_board_review_rows.csv"),
            False,
            "review-only rookie scouting context",
            "not final draft recommendations or private value tuning",
        ),
        (
            "Model v4 pick decision lab",
            Path("local_exports/model_v4/rookie_pick_decision_lab/latest/pick_decision_rows.csv"),
            False,
            "pick-card context and no-baseline receipt for 2026 5.04",
            "not final pick recommendations or exact trade equivalence",
        ),
    )
    rows: list[dict[str, object]] = []
    for area, path, required_for_legal, allowed, blocked in source_specs:
        exists = path.exists()
        count = _csv_count(path) if exists and path.suffix.lower() == ".csv" else 0
        ready = "loaded" if exists else ("missing_required_for_legal_pool" if required_for_legal else "missing_optional_or_inactive")
        if exists and count == 0:
            ready = "loaded_empty"
        rows.append(
            {
                "source_area": area,
                "source_path": str(path),
                "exists": str(exists).lower(),
                "rows": count,
                "readiness_status": ready,
                "allowed_use": allowed,
                "blocked_use": blocked,
                "notes": _readiness_note(area, exists, count),
            }
        )
    return rows


def build_scouting_prep_pool_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    rookie_review = _read_csv_rows(
        Path("local_exports/model_v4/rookie_draft_review/latest/rookie_draft_board_review_rows.csv")
    )
    for row in rookie_review:
        rows.append(
            {
                "player": row.get("prospect_name", ""),
                "position": row.get("position", ""),
                "nfl_team": row.get("nfl_team", ""),
                "college_team": row.get("college", ""),
                "age": "",
                "source_type": "rookie",
                "draftable_status": "rookie_draftable",
                "legal_draftable": "true",
                "roster_owner": "",
                "protected_status": "not_in_protected_roster_source",
                "rookie_status": "incoming_rookie_source_supported",
                "free_agent_status": "",
                "dropped_veteran_status": "",
                "manual_status": "",
                "nwr_draft_value": row.get("league_format_adjusted_score", ""),
                "nwr_rookie_score": row.get("prospect_private_value_review_score", ""),
                "nwr_dynasty_score": "",
                "expected_league_slot_context": "",
                "nwr_slot_context": row.get("draft_board_band", ""),
                "value_threshold_context": "",
                "prospect_talent_context": row.get("evidence_status", ""),
                "landing_spot_context": row.get("nfl_team", ""),
                "draft_capital_context": "",
                "role_path_context": row.get("roster_fit_context", ""),
                "trust_status": _trust_from_confidence(row.get("confidence_cap")),
                "warning_flags": row.get("warning_flags", ""),
                "data_needed": _data_needed_from_warnings(row.get("warning_flags", "")),
                "allowed_use": row.get("allowed_use", ""),
                "blocked_use": row.get("blocked_use", ""),
                "source_path": "local_exports/model_v4/rookie_draft_review/latest/rookie_draft_board_review_rows.csv",
                "source_column": "league_format_adjusted_score",
                "lineage_class": "review_v4_rookie_draft_context",
            }
        )
    free_agents = _read_csv_rows(DRAFT_POOL_PREVIEW_PACK / "fact_available_veterans.csv")
    for row in free_agents:
        position = str(row.get("position") or "")
        if position not in {"QB", "RB", "WR", "TE", "K"}:
            continue
        rows.append(
            {
                "player": row.get("player_name", ""),
                "position": position,
                "nfl_team": row.get("nfl_team", ""),
                "college_team": "",
                "age": "",
                "source_type": "free_agent",
                "draftable_status": "free_agent_draftable",
                "legal_draftable": "true",
                "roster_owner": "",
                "protected_status": "excluded_from_protected_roster_source",
                "rookie_status": "",
                "free_agent_status": row.get("availability_status", ""),
                "dropped_veteran_status": "not_a_dropped_veteran_source",
                "manual_status": "",
                "nwr_draft_value": row.get("draft_value", ""),
                "nwr_rookie_score": "",
                "nwr_dynasty_score": "",
                "expected_league_slot_context": "",
                "nwr_slot_context": row.get("recommended_range", ""),
                "value_threshold_context": row.get("do_not_draft_before_pick", ""),
                "prospect_talent_context": "",
                "landing_spot_context": row.get("nfl_team", ""),
                "draft_capital_context": "",
                "role_path_context": row.get("why_available", ""),
                "trust_status": "source_limited_context",
                "warning_flags": "preview_pack_not_active|free_agent_source_review",
                "data_needed": "Need final legal draft pool confirmation after Roster Declaration Day.",
                "allowed_use": "scouting_prep_context_only",
                "blocked_use": "do_not_use_as_private_value_or_final_draft_recommendation",
                "source_path": str(DRAFT_POOL_PREVIEW_PACK / "fact_available_veterans.csv"),
                "source_column": "draft_value",
                "lineage_class": "preview_free_agent_draftable_source",
            }
        )
    return rows


def _normalize_2025_workbook(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, object]] = []
    for row_idx in range(4, ws.max_row + 1):
        expected_slot = _clean(ws.cell(row_idx, 1).value)
        mock_player = _clean(ws.cell(row_idx, 2).value)
        if expected_slot and mock_player:
            round_label, pick = _parse_round_pick(expected_slot)
            rows.append(
                _history_row(
                    source_file=path.name,
                    source_type="xlsx_user_sheet",
                    draft_year="2025",
                    row_source_area="mock_board",
                    round=round_label,
                    pick=pick,
                    draft_slot_label=expected_slot,
                    expected_slot=expected_slot,
                    player=mock_player,
                    user_note=_clean(ws.cell(row_idx, 3).value),
                    highlight_color_context=_fill_context(ws, row_idx, range(1, 4)),
                    player_type="unknown",
                )
            )
        name = _clean(ws.cell(row_idx, 7).value)
        if name:
            drafted_at = _clean(ws.cell(row_idx, 5).value)
            round_label, pick = _parse_round_pick(drafted_at)
            color_context = _fill_context(ws, row_idx, range(5, 12))
            rows.append(
                _history_row(
                    source_file=path.name,
                    source_type="xlsx_user_sheet",
                    draft_year="2025",
                    row_source_area="user_rank_board",
                    round=round_label,
                    pick=pick,
                    draft_slot_label=drafted_at,
                    drafted_at=drafted_at,
                    user_rank=_clean(ws.cell(row_idx, 6).value),
                    player=name,
                    position=_clean(ws.cell(row_idx, 9).value).upper(),
                    nfl_team_at_draft=_clean(ws.cell(row_idx, 8).value),
                    draft_capital=_clean(ws.cell(row_idx, 10).value),
                    age=_clean(ws.cell(row_idx, 11).value),
                    user_note=_clean(ws.cell(row_idx, 12).value),
                    user_drafted_flag=str(_norm_name(name) in EXPLICIT_2025_USER_DRAFTED).lower(),
                    user_must_draft_at_cost_flag=str("yellow" in color_context).lower(),
                    highlight_color_context=color_context,
                    player_type=_player_type_from_fields(
                        draft_capital=_clean(ws.cell(row_idx, 10).value),
                        age=_clean(ws.cell(row_idx, 11).value),
                    ),
                )
            )
    return rows


def _normalize_2024_workbook(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, object]] = []
    for row_idx in range(1, ws.max_row + 1):
        slot = _clean(ws.cell(row_idx, 1).value)
        ranked_name = _clean(ws.cell(row_idx, 2).value)
        if slot and ranked_name:
            rank, player = _split_ranked_name(ranked_name)
            round_label, pick = _parse_round_pick(slot)
            rows.append(
                _history_row(
                    source_file=path.name,
                    source_type="xlsx_user_sheet",
                    draft_year="2024",
                    row_source_area="mock_board",
                    round=round_label,
                    pick=pick,
                    draft_slot_label=slot,
                    expected_slot=slot,
                    user_rank=rank,
                    player=player,
                    age=_clean(ws.cell(row_idx, 3).value),
                    user_note=_clean(ws.cell(row_idx, 5).value),
                    highlight_color_context=_fill_context(ws, row_idx, range(1, 6)),
                    player_type="unknown",
                )
            )
        right = _clean(ws.cell(row_idx, 6).value)
        if right:
            rank, player = _split_ranked_name(right)
            drafted_at = _clean(ws.cell(row_idx, 7).value)
            round_label, pick = _parse_round_pick(drafted_at)
            rows.append(
                _history_row(
                    source_file=path.name,
                    source_type="xlsx_user_sheet",
                    draft_year="2024",
                    row_source_area="user_rank_board",
                    round=round_label,
                    pick=pick,
                    draft_slot_label=drafted_at,
                    drafted_at=drafted_at,
                    user_rank=rank,
                    player=player,
                    draft_capital=drafted_at,
                    user_note=_clean(ws.cell(row_idx, 8).value),
                    highlight_color_context=_fill_context(ws, row_idx, range(6, 9)),
                    player_type="rookie",
                )
            )
    return rows


def _normalize_generic_workbook(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    rows: list[dict[str, object]] = []
    draft_year = _year_from_name(path.name)
    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            values = [_clean(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
            if not any(values):
                continue
            rows.append(
                _history_row(
                    source_file=path.name,
                    source_type="xlsx_user_sheet",
                    draft_year=draft_year,
                    row_source_area="unknown",
                    player=" | ".join(value for value in values if value),
                    warning_flags="generic_workbook_row_not_structured",
                    data_needed="Review workbook row manually before using as historical context.",
                )
            )
    return rows


def _normalize_pdf_transcription(path: Path) -> list[dict[str, object]]:
    rows = []
    for row in _read_csv_rows(path):
        confidence = str(row.get("parse_confidence") or "").lower()
        draft_slot = row.get("draft_slot_label") or ""
        round_label = row.get("round") or _parse_round_pick(draft_slot)[0]
        pick = row.get("pick_in_round") or _parse_round_pick(draft_slot)[1]
        rows.append(
            _history_row(
                source_file=row.get("source_file") or path.name,
                source_type="pdf_best_effort_transcription",
                draft_year=row.get("draft_year", ""),
                row_source_area="actual_draft",
                round=round_label,
                pick=pick,
                overall_pick=row.get("overall_pick", ""),
                draft_slot_label=draft_slot,
                drafted_at=draft_slot,
                drafting_team=row.get("drafting_team_or_slot_owner", ""),
                player=row.get("normalized_player_guess") or row.get("player_raw", ""),
                transcription_confidence=confidence,
                allowed_use=row.get("allowed_use", ""),
                blocked_use=row.get("blocked_use", ""),
                warning_flags="low_confidence_pdf_transcription" if confidence == "low" else "",
                data_needed=(
                    "Verify against included page image before relying on this row."
                    if confidence == "low"
                    else ""
                ),
                player_type="unknown",
            )
        )
    return rows


def _history_row(**kwargs: object) -> dict[str, object]:
    row = {key: "" for key in PRIOR_HISTORY_HEADER}
    row.update(kwargs)
    row["player"] = _clean(row.get("player"))
    row["position"] = _clean(row.get("position")).upper()
    row["normalized_player_key"] = _norm_name(str(row.get("player") or ""))
    row["user_drafted_flag"] = str(row.get("user_drafted_flag") or "false").lower()
    row["user_must_draft_at_cost_flag"] = str(row.get("user_must_draft_at_cost_flag") or "false").lower()
    row["allowed_use"] = row.get("allowed_use") or (
        "historical league draft behavior; expected pick range context; "
        "display-only draft prep context; user preference/context notes"
    )
    row["blocked_use"] = row.get("blocked_use") or (
        "NWR private value; NWR Draft Value; NWR Dynasty Score; pick baselines; "
        "VORP; replacement; final draft recommendations"
    )
    return row


def _write_current_state_audit(path: Path, readiness_rows: list[dict[str, object]]) -> None:
    active_rookies = _readiness_by_area(readiness_rows, "active confirmed legal draftable pool")
    active_fa = _readiness_by_area(readiness_rows, "active available veterans/free agents")
    preview_rookies = _readiness_by_area(readiness_rows, "preview rookie draftables")
    preview_fa = _readiness_by_area(readiness_rows, "preview free agents")
    content = f"""# Draft Prep Current State Audit - 2026-06-09

## Current Draft Board files
- Page: `app/pages/06_draft_board.py`.
- UI/session helper: `app/components/draft_session.py`.
- Product contract/ranking table: `src/services/draft_ux_service.py`.
- Core draft board builder: `src/services/draft_service.py`.
- Live/mock board state: `src/services/draft_state_service.py`.
- Saved mock drafts: `src/services/mock_draft_storage_service.py`.
- Draftable preview builder: `src/services/real_draft_pool_preview_service.py`.
- Rookie review exports: `src/services/model_v4_sprint14e_rookie_draft_review_service.py`.

## Current behavior
- The page currently mixes Draft Board planning, draft rankings, pick grid, live/mock drafted state, save/load mock controls, rookie model receipts, source-risk expanders, and audit/debug context.
- The scheduled draft grid is built from `fact_future_picks.csv`; the active pack has 50 scheduled picks for a 5 round x 10 team offline draft.
- Niners picks are resolved by matching `current_team_name`/team id to the `team_id='niners'` alias in `build_draft_room`.
- Mock/live state lives in Streamlit session state through `app/components/draft_session.py` and can be saved as JSON under `local_exports/mock_drafts`.

## Draftable pool status
- Active pack rookie draftables: {active_rookies.get('readiness_status')} ({active_rookies.get('rows')} rows).
- Active pack free agents/veterans: {active_fa.get('readiness_status')} ({active_fa.get('rows')} rows).
- Preview pack rookie draftables: {preview_rookies.get('readiness_status')} ({preview_rookies.get('rows')} rows).
- Preview pack free agents: {preview_fa.get('readiness_status')} ({preview_fa.get('rows')} rows).
- The current active pack draftable pool is effectively 0 because `fact_rookie_draftables.csv`, `fact_available_veterans.csv`, and `fact_manual_draftables.csv` are not present in the active pack.
- Draft-pool preview packs exist and can support scouting/readiness, but this task does not promote them into active app state.

## Model v4 rookie/pick exports loaded today
- `local_exports/model_v4/rookie_draft_review/latest/rookie_draft_board_review_rows.csv`
- `local_exports/model_v4/rookie_draft_review/latest/rookie_pick_candidate_review_rows.csv`
- `local_exports/model_v4/rookie_pick_decision_lab/latest/pick_decision_rows.csv`
- `local_exports/model_v4/human_decision_review_prep/latest/pick_review_cards.csv`
- These are review/product-prep context, not final recommendations.

## 2026 5.04
- The active pack contains an old `fact_pick_values.csv` row for `2026 5.04`, but Model v4 pick-decision rows correctly label it `manual_only_no_exact_model_baseline`.
- Draft Prep must keep `2026 5.04` visible as a late-round pick card with `No Baseline` / `Manual Watchlist` language.

## Missing legal source
- Dropped/released veterans are expected after Roster Declaration Day.
- Missing dropped-player source must not be interpreted as proof that no veterans are draftable.
"""
    path.write_text(content, encoding="utf-8")


def _write_prior_normalization_doc(path: Path, rows: list[dict[str, object]], package: Path) -> None:
    by_source = Counter(str(row["source_file"]) for row in rows)
    drafted = [
        row["player"]
        for row in rows
        if row.get("draft_year") == "2025" and row.get("user_drafted_flag") == "true"
    ]
    yellow = sum(1 for row in rows if row.get("user_must_draft_at_cost_flag") == "true")
    low_pdf = sum(1 for row in rows if row.get("transcription_confidence") == "low")
    content = f"""# Prior Draft History Normalization - 2026-06-09

## Package
- Source root: `{package}`
- Normalized CSV: `local_exports/model_v4/draft_prep/latest/prior_league_draft_history_review_rows.csv`

## Row counts by source
{_markdown_count_table(by_source)}

## 2025 workbook handling
- The left side is preserved as `mock_board`.
- The right side is preserved as `user_rank_board`.
- `Drafted at`, rank, player, NFL team, position, draft capital, age, notes, and visible fill context are preserved where present.
- Explicit 2025 user drafted list is ground truth: {", ".join(drafted)}.
- Green fill is retained only as `highlight_color_context`; it is not used as drafted ground truth.
- Yellow fill is interpreted as `user_must_draft_at_cost_flag=true`, meaning high-interest at cost, not draft at any cost.
- Yellow/highlight rows found: {yellow}.

## PDF handling
- The image-only league-manager PDF is represented through `drafts_pdf_official_transcription_best_effort.csv`.
- Low-confidence PDF rows found: {low_pdf}.
- Low-confidence rows carry `data_needed=Verify against included page image before relying on this row.`
- No unreadable picks were invented.

## Guardrails
- Prior draft history, spreadsheet highlights, and user notes are display-only/context-only.
- They are blocked from NWR private value, NWR Draft Value, pick baselines, VORP, replacement, and final draft recommendations.
"""
    path.write_text(content, encoding="utf-8")


def _write_behavior_doc(path: Path, behavior_rows: list[dict[str, object]], prior_rows: list[dict[str, object]]) -> None:
    first_round_vets = sum(_safe_int(row["first_round_veteran_count"]) for row in behavior_rows)
    first_round_rookies = sum(_safe_int(row["first_round_rookie_count"]) for row in behavior_rows)
    round_one_rows = [row for row in prior_rows if str(row.get("round")) == "1"]
    source_limits = sum(1 for row in prior_rows if row.get("transcription_confidence") == "low")
    content = f"""# Prior League Draft Behavior Summary - 2026-06-09

## Outputs
- CSV: `local_exports/model_v4/draft_prep/latest/prior_league_draft_behavior_summary.csv`

## Summary
- Round 1 rows with usable player-type context are mixed, but the PDF rows often lack position/player-type evidence.
- First-round rookie rows counted from structured sheets: {first_round_rookies}.
- First-round veteran/free-agent rows counted from structured sheets: {first_round_vets}.
- Total Round 1 context rows: {len(round_one_rows)}.
- Low-confidence PDF transcription rows needing verification: {source_limits}.

## What history supports
- Draft Prep should show expected league range as context because prior sheets/PDFs are pick-order oriented.
- Historical behavior supports a mixed-pool mindset: rookies plus veterans/free agents, especially in early rounds.
- Position tendencies by round can be summarized when position exists, but older PDF rows need verification before strong claims.
- QBs/TEs should be visible as league-history context, not used to modify private value.

## 2025 product lesson
- The 2025 workbook is a better UI/product reference than the current audit cockpit.
- It naturally separates expected/mock board from a user-ranked review board and includes notes, cost, draft capital, age, and actual drafted-at context.
- Draft Prep should translate that into pick cards and candidate windows, not another raw rookie ranking table.

## Guardrail
- This summary is historical behavior context only and is blocked from private scoring.
"""
    path.write_text(content, encoding="utf-8")


def _write_source_contract(path: Path) -> None:
    content = """# Draftable Pool Source Contract - 2026-06-09

## Legal draftable source types
1. `rookie`: incoming rookies from admitted rookie/prospect sources.
2. `free_agent`: unrostered/legal free agents from a current source.
3. `dropped_veteran`: officially released/dropped veterans after Roster Declaration Day.
4. `manual`: manually added draftable players with explicit source notes.
5. `protected_not_draftable`: protected roster players, blocked from confirmed draftable status.

## Draftable statuses
- `confirmed_draftable`
- `rookie_draftable`
- `free_agent_draftable`
- `dropped_veteran_draftable`
- `manual_draftable`
- `scouting_only`
- `protected_not_draftable`
- `needs_source`
- `no_baseline`
- `hidden_kicker`

## Source precedence
1. Official dropped/released veteran list after Roster Declaration Day.
2. Current rookie draftable/prospect source.
3. Current free-agent/unrostered source.
4. Manual draftable additions.
5. Protected roster source as a blocklist, not as draftable evidence.

## Blocked from private value
Market rank, league rank, ADP, startup, projections, consensus, public ranks, trade calculators, RotoWire projections/rankings, prior draft history, spreadsheet highlights, and legacy active-pack scores cannot create or modify NWR Draft Value or NWR Dynasty Score.
"""
    path.write_text(content, encoding="utf-8")


def _write_readiness_doc(
    path: Path,
    readiness_rows: list[dict[str, object]],
    scouting_rows: list[dict[str, object]],
) -> None:
    active_missing = [
        row for row in readiness_rows if row["source_area"].startswith("active") and row["exists"] == "false"
    ]
    dropped = _readiness_by_area(readiness_rows, "dropped/released veterans")
    rookie_rows = _readiness_by_area(readiness_rows, "Model v4 rookie review")
    preview_fa = _readiness_by_area(readiness_rows, "preview free agents")
    content = f"""# Draftable Pool Source Readiness - 2026-06-09

## Can we build a confirmed legal draftable pool now?
No. The active pack does not contain active draftable pool files, and the official dropped/released veteran source is missing until Roster Declaration Day.

## Can we build a scouting prep board now?
Yes. `scouting_prep_pool_review_rows.csv` was created with {len(scouting_rows)} review rows from existing rookie review and preview free-agent sources.

## Source readback
- Active missing draftable sources: {len(active_missing)}.
- Model v4 rookie review rows: {rookie_rows.get('rows')}.
- Preview free-agent rows: {preview_fa.get('rows')}.
- Dropped/released veteran source: {dropped.get('readiness_status')}.

## Protected roster source
`local_exports/data_packs/lve_sleeper_20260505_pdf_ranks/fact_rosters.csv` is available and should block protected players from being marked confirmed draftable.

## 2026 5.04
Show as a visible late-round pick card only. Use `No Baseline` / `Manual Watchlist`; do not invent exact equivalence.
"""
    path.write_text(content, encoding="utf-8")


def _write_pick_window_spec(path: Path) -> None:
    cards = "\n".join(
        f"- `{label}`: "
        + ("No Baseline / Manual Watchlist / no exact equivalence." if label == "2026 5.04" else "baseline/value status, candidate count, source readiness, and data blockers.")
        for label in OWNED_PICK_LABELS
    )
    content = f"""# Draft Prep Pick Window Spec - 2026-06-09

## Pick cards
{cards}

## Candidate labels
- In Range
- Likely Gone
- Value If Falls
- Possible Reach
- Expensive vs NWR
- Favorite At Cost
- Must Review At Cost
- Needs Scouting
- Source Limited
- Manual Watchlist
- No Baseline

## Default future table columns
- Fit Band
- Expected League Slot
- NWR Slot
- Value Threshold
- Player
- Pos
- NFL Team
- Source Type
- NWR Draft Value
- Prospect Talent
- Landing Spot
- Draft Capital
- Role Path
- Trust
- Warnings
- Data Needed
- Draftable Status

## Guardrails
Pick-vs-player comparisons are advanced/context-only. `2026 5.04` must never receive an invented baseline or exact equivalence.
"""
    path.write_text(content, encoding="utf-8")


def _write_page_architecture(path: Path) -> None:
    content = """# Draft Prep Page Architecture - 2026-06-09

## Draft Prep
- Pick-by-pick planning.
- Candidate windows around owned picks.
- Draftable pool readiness.
- Scouting queue.
- League-history context.
- Conservative, cost-aware labels only.

## Live Draft Room
- Mock/live draft grid.
- Drafted player state.
- Undo/replace/reset/save/load.
- Best remaining during the draft.
- This task does not implement this split.

## Advanced Audit
- Receipts.
- Historical comps.
- Startup slot context.
- Raw warning tables.
- Source risk heatmap.
- Pick equivalence/debug context.

## Recommendation
Keep Draft Prep focused on planning and move all live/mock state to a future Live Draft Room page.
"""
    path.write_text(content, encoding="utf-8")


def _fill_context(ws, row_idx: int, columns: Iterable[int]) -> str:
    labels: list[str] = []
    for col in columns:
        cell = ws.cell(row_idx, col)
        if not cell.fill or not cell.fill.fill_type:
            continue
        rgb = str(cell.fill.fgColor.rgb or "").upper()
        if rgb == "FFFFFF00":
            labels.append(f"{cell.coordinate}:yellow_must_draft_at_cost")
        elif rgb == "FF00FF00":
            labels.append(f"{cell.coordinate}:green_user_drafted_context_only")
        elif rgb and rgb not in {"00000000", "FFFFFFFF"}:
            labels.append(f"{cell.coordinate}:fill_{rgb}")
    return "|".join(labels)


def _player_type_from_fields(*, draft_capital: str, age: str) -> str:
    if draft_capital:
        return "rookie"
    age_value = _safe_float(age)
    if age_value is not None and age_value >= 24.5:
        return "veteran_or_free_agent"
    return "unknown"


def _trust_from_confidence(raw: object) -> str:
    value = _safe_float(raw)
    if value is None:
        return "source_limited"
    if value < 0.75:
        return "source_limited"
    if value < 0.9:
        return "usable_with_confidence_cap"
    return "review_context"


def _data_needed_from_warnings(raw: object) -> str:
    text = str(raw or "")
    needed: list[str] = []
    if "missing" in text:
        needed.append("Need missing source evidence review.")
    if "source_limited" in text or "third_party" in text:
        needed.append("Source-limited evidence; verify before relying on candidate window.")
    if "team_mismatch" in text or "quarantined" in text:
        needed.append("Need identity/team/source mapping review.")
    return " ".join(needed)


def _readiness_note(area: str, exists: bool, count: int) -> str:
    if area == "dropped/released veterans":
        return "Expected after Roster Declaration Day; currently missing."
    if not exists:
        return "Missing in this source lane."
    if count == 0:
        return "File exists but has zero rows."
    return "Loaded for review/context."


def _source_ready(rows: list[dict[str, object]], area: str) -> bool:
    row = _readiness_by_area(rows, area)
    return row.get("readiness_status") == "loaded"


def _readiness_by_area(rows: list[dict[str, object]], area: str) -> dict[str, object]:
    return next((row for row in rows if row.get("source_area") == area), {})


def _split_ranked_name(value: str) -> tuple[str, str]:
    match = re.match(r"^\s*(\d+)\s+(.+?)\s*$", value)
    if match:
        return match.group(1), match.group(2).strip()
    return "", value


def _parse_round_pick(value: object) -> tuple[str, str]:
    text = _clean(value)
    match = re.search(r"(\d+)\.(\d+)", text)
    if not match:
        return "", ""
    pick_text = match.group(2)
    pick = 10 if len(pick_text) == 1 else int(pick_text)
    return str(int(match.group(1))), str(pick)


def _year_from_name(name: str) -> str:
    match = re.search(r"(20\d{2})", name)
    return match.group(1) if match else ""


def _norm_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def _clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _safe_int(value: object) -> int:
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return 0


def _safe_float(value: object) -> float | None:
    try:
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def _csv_count(path: Path) -> int:
    return len(_read_csv_rows(path))


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, header: Iterable[str], rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(header), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _markdown_count_table(counter: Counter[str]) -> str:
    lines = ["| Source | Rows |", "| --- | ---: |"]
    for key, value in sorted(counter.items()):
        lines.append(f"| `{key}` | {value} |")
    return "\n".join(lines)
